import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import datetime
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="TPCRA v3.0 Dashboard",
    page_icon="🔐",
    layout="wide",
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
.kpi-label { font-size: 11px; font-weight: 600; text-transform: uppercase;
             letter-spacing: 0.07em; color: #6c757d; margin-bottom: 4px; }
.kpi-value { font-size: 28px; font-weight: 600; line-height: 1.1; }
.tier-critical { color: #A32D2D; }
.tier-high     { color: #854F0B; }
.tier-medium   { color: #185FA5; }
.tier-low      { color: #3B6D11; }
.response-pill {
    display: inline-block; padding: 2px 10px; border-radius: 20px;
    font-size: 11px; font-weight: 600; letter-spacing: 0.03em;
}
.stTabs [data-baseweb="tab"] { font-size: 13px; padding: 8px 16px; }
div[data-testid="metric-container"] {
    background: #f8f9fa; border-radius: 10px; padding: 14px;
    border: 1px solid #e9ecef;
}
</style>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
DOMAIN_MAP = {
    "A": "Organizational Management",
    "B": "Human Resource Management",
    "C": "Infrastructure Security",
    "D": "Data Protection",
    "E": "Access Management",
    "F": "Application Security",
    "G": "System Security",
    "H": "Email Security",
    "I": "Mobile Devices",
    "J": "Incident Response",
    "K": "Cloud Services",
    "L": "Business Continuity",
    "M": "Supply Chain & Physical Security",
    "N": "AI & Emerging Technology Risk",
}

RESP_COLORS = {
    "Yes":     "#639922",
    "No":      "#E24B4A",
    "Partial": "#EF9F27",
    "N/A":     "#B4B2A9",
}

RESP_PILL = {
    "Yes":     "background:#EAF3DE;color:#27500A",
    "No":      "background:#FCEBEB;color:#A32D2D",
    "Partial": "background:#FAEEDA;color:#633806",
    "N/A":     "background:#F1EFE8;color:#5F5E5A",
    "—":       "background:#F1EFE8;color:#888780",
}

TIER_PILL = {
    "Critical": "background:#FCEBEB;color:#A32D2D",
    "High":     "background:#FAEEDA;color:#854F0B",
    "Medium":   "background:#E6F1FB;color:#0C447C",
    "Low":      "background:#EAF3DE;color:#3B6D11",
}

RATING_THRESHOLDS = [
    (90, "✅ Low",      "#639922"),
    (70, "🟡 Medium",   "#BA7517"),
    (50, "🟠 High",     "#D85A30"),
    (0,  "🔴 Critical", "#A32D2D"),
]

YES_VALS  = {"yes", "y"}
NO_VALS   = {"no", "n"}
NA_VALS   = {"n/a", "na", "not applicable"}
PART_VALS = {"partial", "partly", "partially"}
EVIDENCE_STATUS = {"submitted", "provided", "received", "complete", "done", "yes"}

# ── Static gap database (from TPCRA v3.0 questionnaire analysis) ───────────────
GAP_DB = [
    {
        "id": "A", "name": "Organizational Management", "risk": "Critical",
        "critical": 3, "high": 4, "medium": 2, "total": 9,
        "gaps": [
            {"ref": "A.1",  "text": "IT Security policies not formally established or documented", "tier": "Critical"},
            {"ref": "A.4e", "text": "Security incident response not covered in policy", "tier": "Critical"},
            {"ref": "A.6",  "text": "Policy not aligned to regulatory requirements / customer data protection", "tier": "Critical"},
            {"ref": "A.5",  "text": "Policies not formally acknowledged / signed by all staff", "tier": "High"},
            {"ref": "A.7",  "text": "No formal risk management framework (ISO 27001, NIST CSF)", "tier": "High"},
            {"ref": "A.8",  "text": "No independent security certification or audit report", "tier": "High"},
        ],
        "recs": [
            "Establish and approve a formal Information Security Policy covering all required topics (A.4a–k) within 30 days",
            "Align policies to applicable regulations (DPA, GDPR, PDPA) and obtain senior leadership sign-off",
            "Initiate ISO 27001 or NIST CSF gap assessment; target certification within 12 months",
            "Obtain and share current third-party security audit report or certification",
        ],
    },
    {
        "id": "B", "name": "Human Resource Management", "risk": "Critical",
        "critical": 2, "high": 2, "medium": 1, "total": 6,
        "gaps": [
            {"ref": "B.3", "text": "Background screening and NDA not enforced for new hires / contractors", "tier": "Critical"},
            {"ref": "B.5", "text": "No formal offboarding process to revoke access on last working day", "tier": "Critical"},
            {"ref": "B.2", "text": "Security awareness training not extended to all employees and contractors", "tier": "High"},
            {"ref": "B.6", "text": "No insider threat controls (UBA, session monitoring, separation of duties)", "tier": "High"},
        ],
        "recs": [
            "Implement mandatory background checks and NDA execution before granting system access",
            "Formalize offboarding checklist to revoke all logical and physical access on day of departure",
            "Roll out annual security awareness training to 100% of staff, including third-party contractors",
            "Deploy user behaviour analytics or privileged session monitoring to detect insider threats",
        ],
    },
    {
        "id": "C", "name": "Infrastructure Security", "risk": "Critical",
        "critical": 11, "high": 7, "medium": 0, "total": 20,
        "gaps": [
            {"ref": "C.1.2", "text": "No clear network segmentation; externally-facing apps not in DMZ", "tier": "Critical"},
            {"ref": "C.1.3", "text": "Corporate network not separated from internet by firewall", "tier": "Critical"},
            {"ref": "C.1.5", "text": "Inter-system communication not using encrypted protocols only", "tier": "Critical"},
            {"ref": "C.1.6", "text": "Deprecated protocols (TLS 1.0/1.1, Telnet, FTP) not disabled", "tier": "Critical"},
            {"ref": "C.2.1", "text": "IPS/firewall not blocking unauthorized external access", "tier": "Critical"},
            {"ref": "C.2.4", "text": "WAF not deployed on internet-facing applications", "tier": "Critical"},
            {"ref": "C.5.1", "text": "Remote connections do not require MFA and encrypted channel", "tier": "Critical"},
            {"ref": "C.5.4", "text": "Remote access not revoked promptly (within 24 hours) when no longer required", "tier": "Critical"},
            {"ref": "C.6.2", "text": "No defined SLA timelines for applying security patches by criticality", "tier": "Critical"},
            {"ref": "C.3.5", "text": "No centralized security event monitoring (SIEM/SOC)", "tier": "Critical"},
        ],
        "recs": [
            "Implement network segmentation with DMZ for all internet-facing services; firewall all public access points",
            "Disable TLS 1.0/1.1, SSLv3, FTP, Telnet, and enforce TLS 1.2+ across all systems immediately",
            "Deploy WAF and enforce MFA for all remote and cloud console access",
            "Define and publish patch SLA: Critical ≤24h, High ≤7 days, Medium ≤30 days",
            "Deploy SIEM or engage SOC for centralized log aggregation and alerting",
        ],
    },
    {
        "id": "D", "name": "Data Protection", "risk": "Critical",
        "critical": 8, "high": 4, "medium": 1, "total": 14,
        "gaps": [
            {"ref": "D.1.3",  "text": "No technical isolation of client data from other tenants", "tier": "Critical"},
            {"ref": "D.1.4",  "text": "Data not protected by encryption and access controls across lifecycle", "tier": "Critical"},
            {"ref": "D.1.5",  "text": "No process for secure data destruction on contract termination", "tier": "Critical"},
            {"ref": "D.1.6",  "text": "Controls missing to prevent unauthorized extraction of customer data", "tier": "Critical"},
            {"ref": "D.2.1",  "text": "Strong cryptographic protocols (TLS 1.2+) not protecting data in transit", "tier": "Critical"},
            {"ref": "D.2.3",  "text": "Cryptographic key management lacks dual-control and secure storage", "tier": "Critical"},
            {"ref": "D.2.4",  "text": "Backup data not encrypted using industry-accepted standards", "tier": "Critical"},
            {"ref": "D.2.5",  "text": "Deprecated cipher suites (MD5, SHA-1, RC4) not identified and disabled", "tier": "Critical"},
        ],
        "recs": [
            "Implement data classification tagging and enforce access controls aligned to sensitivity level",
            "Encrypt all data at rest (AES-256) and in transit (TLS 1.2+); include backup media",
            "Establish a formal key management lifecycle policy with dual-control and secure vaulting",
            "Define and enforce data retention and secure destruction procedures, including for backups",
            "Deploy DLP solution to detect and block unauthorized data exfiltration",
        ],
    },
    {
        "id": "E", "name": "Access Management", "risk": "Critical",
        "critical": 8, "high": 7, "medium": 2, "total": 18,
        "gaps": [
            {"ref": "E.1.1",  "text": "Access to critical systems not gated by formal approved request process", "tier": "Critical"},
            {"ref": "E.1.3",  "text": "No periodic review of user access to systems and databases", "tier": "Critical"},
            {"ref": "E.1.7",  "text": "Shared user IDs or privileged accounts in use", "tier": "Critical"},
            {"ref": "E.1.8",  "text": "MFA not enforced for remote, cloud, and privileged access", "tier": "Critical"},
            {"ref": "E.2.3",  "text": "MFA not enforced for all user accounts", "tier": "Critical"},
            {"ref": "E.2.12", "text": "Passwords not stored with strong salted hashing (bcrypt/Argon2)", "tier": "Critical"},
            {"ref": "E.3.1",  "text": "No documented procedures for privileged / break-glass access", "tier": "Critical"},
            {"ref": "E.3.2",  "text": "Privileged access not reviewed quarterly with dual-control / JIT", "tier": "Critical"},
        ],
        "recs": [
            "Enforce MFA organisation-wide — prioritise privileged, remote, and cloud console access",
            "Implement quarterly user access reviews with automated provisioning/deprovisioning workflows",
            "Eliminate shared accounts; enforce unique user IDs and least-privilege principles",
            "Deploy or evaluate PAM solution for privileged credential vaulting and session recording",
            "Migrate to bcrypt or Argon2 password hashing; audit credential storage immediately",
        ],
    },
    {
        "id": "F", "name": "Application Security", "risk": "Critical",
        "critical": 4, "high": 4, "medium": 1, "total": 10,
        "gaps": [
            {"ref": "F.3", "text": "No security / pentest conducted before application onboarding", "tier": "Critical"},
            {"ref": "F.4", "text": "Secure coding guidelines (OWASP Top 10) not formally adopted", "tier": "Critical"},
            {"ref": "F.9", "text": "API security controls (OAuth, rate-limiting, input validation) not implemented", "tier": "Critical"},
            {"ref": "F.6", "text": "Security requirements not formally embedded in SDLC", "tier": "High"},
            {"ref": "F.7", "text": "No SCA process for open-source / third-party library vulnerabilities", "tier": "High"},
        ],
        "recs": [
            "Mandate penetration testing and secure code review before any application goes live",
            "Adopt OWASP Top 10 and SANS CWE Top 25 as formal secure coding standards",
            "Implement API gateway with OAuth 2.0, rate limiting, and input validation",
            "Integrate SAST/DAST tooling and SCA into CI/CD pipelines",
        ],
    },
    {
        "id": "G", "name": "System Security", "risk": "Critical",
        "critical": 6, "high": 4, "medium": 0, "total": 11,
        "gaps": [
            {"ref": "G.2.1", "text": "No timely review of security audit logs; anomalies not acted on", "tier": "Critical"},
            {"ref": "G.2.2", "text": "Automated audit trails not capturing admin actions and auth changes", "tier": "Critical"},
            {"ref": "G.2.3", "text": "Security logs not reviewed for anomalies via SIEM correlation", "tier": "Critical"},
            {"ref": "G.2.4", "text": "Logging facilities not protected against tampering / unauthorized modification", "tier": "Critical"},
            {"ref": "G.3.1", "text": "No real-time monitoring of security events on critical systems", "tier": "Critical"},
            {"ref": "G.3.5", "text": "No quarterly automated vulnerability scanning with defined SLA remediation", "tier": "Critical"},
        ],
        "recs": [
            "Deploy SIEM with automated correlation rules; assign SOC team / alerting coverage 24×7",
            "Implement tamper-protected, write-once log storage with minimum 12-month retention",
            "Schedule quarterly automated vulnerability scans with SLA-tracked remediation",
            "Enable full audit trails for admin actions, authentication events, and access changes",
        ],
    },
    {
        "id": "H", "name": "Email Security", "risk": "High",
        "critical": 1, "high": 4, "medium": 0, "total": 5,
        "gaps": [
            {"ref": "H.3", "text": "Email attachments and URLs not scanned for malware/phishing (sandboxing)", "tier": "Critical"},
            {"ref": "H.5", "text": "SPF, DKIM, DMARC email authentication protocols not implemented", "tier": "High"},
            {"ref": "H.2", "text": "Email data confidentiality and integrity not protected by technical controls", "tier": "High"},
            {"ref": "H.6", "text": "Simulated phishing exercises not conducted at least annually", "tier": "High"},
        ],
        "recs": [
            "Deploy email gateway with sandboxing for URL and attachment analysis",
            "Implement SPF, DKIM, and DMARC with at minimum quarantine policy",
            "Conduct simulated phishing exercises at least annually; track click rates",
        ],
    },
    {
        "id": "I", "name": "Mobile Devices", "risk": "High",
        "critical": 1, "high": 4, "medium": 0, "total": 5,
        "gaps": [
            {"ref": "I.3", "text": "Laptops do not enforce full-disk encryption (BitLocker / FileVault)", "tier": "Critical"},
            {"ref": "I.2", "text": "Mobile devices not enrolled in MDM with security software / PIN enforced", "tier": "High"},
            {"ref": "I.5", "text": "BYOD devices lack corporate data isolation and remote wipe capability", "tier": "High"},
            {"ref": "I.6", "text": "Jailbroken/rooted devices not automatically blocked from corporate access", "tier": "High"},
        ],
        "recs": [
            "Enforce full-disk encryption on all laptops via Group Policy or MDM",
            "Enrol all corporate and BYOD devices into MDM; enforce PIN, remote wipe, and compliance check",
            "Block jailbroken/rooted devices via MDM compliance policy",
        ],
    },
    {
        "id": "J", "name": "Incident Response", "risk": "Critical",
        "critical": 3, "high": 3, "medium": 1, "total": 7,
        "gaps": [
            {"ref": "J.2", "text": "No documented Incident Response Plan with roles and escalation paths", "tier": "Critical"},
            {"ref": "J.3", "text": "IRP not tested at least annually (tabletop / simulation)", "tier": "Critical"},
            {"ref": "J.5", "text": "No defined contractual SLA for notifying client of security incidents", "tier": "Critical"},
            {"ref": "J.4", "text": "IRP not reviewed and updated at least annually", "tier": "High"},
            {"ref": "J.7", "text": "No 24×7 security operations capability (in-house or outsourced SOC)", "tier": "High"},
        ],
        "recs": [
            "Develop and approve an Incident Response Plan covering all incident types with escalation matrix",
            "Conduct annual tabletop exercise; document and track lessons learned",
            "Define and agree contractual incident notification SLA (recommend ≤4 hours for confirmed breaches)",
            "Establish or contract 24×7 SOC capability for continuous threat monitoring",
        ],
    },
    {
        "id": "K", "name": "Cloud Services", "risk": "Critical",
        "critical": 7, "high": 3, "medium": 1, "total": 11,
        "gaps": [
            {"ref": "K.2", "text": "CSP lacks third-party security certification (ISO 27001, SOC 2, etc.)", "tier": "Critical"},
            {"ref": "K.4", "text": "No logical or physical data segregation between tenants", "tier": "Critical"},
            {"ref": "K.5", "text": "Hypervisor admin access not restricted with least privilege and MFA", "tier": "Critical"},
            {"ref": "K.6", "text": "Cryptographic key management process not formal (AWS KMS / Azure Key Vault)", "tier": "Critical"},
            {"ref": "K.8", "text": "Security logs not generated, retained, and accessible on demand", "tier": "Critical"},
            {"ref": "K.9", "text": "No confirmed data deletion process with certificate upon termination", "tier": "Critical"},
        ],
        "recs": [
            "Verify CSP holds current ISO 27001, SOC 2 Type 2, or equivalent certification",
            "Confirm tenant data isolation and request Shared Responsibility Matrix from CSP",
            "Enforce MFA and least privilege for all hypervisor and admin console access",
            "Implement formal key management via AWS KMS / Azure Key Vault; document lifecycle",
            "Obtain written data deletion certificate commitment; include in contract",
        ],
    },
    {
        "id": "L", "name": "Business Continuity", "risk": "Critical",
        "critical": 4, "high": 2, "medium": 0, "total": 6,
        "gaps": [
            {"ref": "L.1", "text": "No formally approved Business Continuity Plan (BCP)", "tier": "Critical"},
            {"ref": "L.3", "text": "BCP not tested at least annually", "tier": "Critical"},
            {"ref": "L.4", "text": "RTO and RPO targets not defined or validated through testing", "tier": "Critical"},
            {"ref": "L.5", "text": "Backups not stored off-site or tested for restorability annually", "tier": "Critical"},
        ],
        "recs": [
            "Develop, approve, and communicate a formal BCP with RTO/RPO targets for critical services",
            "Schedule annual BCP/DR tabletop and failover test; document results and remediation actions",
            "Implement off-site or cross-region encrypted backup storage with regular restore tests",
        ],
    },
    {
        "id": "M", "name": "Supply Chain & Physical Security", "risk": "Critical",
        "critical": 3, "high": 5, "medium": 2, "total": 10,
        "gaps": [
            {"ref": "M.1", "text": "No formal TPRM program to assess sub-processors", "tier": "Critical"},
            {"ref": "M.2", "text": "Sub-processors / fourth parties with data access not identified", "tier": "Critical"},
            {"ref": "M.3", "text": "Security requirements not contractually flowed down to sub-processors", "tier": "Critical"},
            {"ref": "M.6", "text": "Physical access to server rooms not restricted with multi-factor controls", "tier": "High"},
            {"ref": "M.9", "text": "Environmental controls (fire suppression, UPS, CCTV) not confirmed for data hosting", "tier": "High"},
        ],
        "recs": [
            "Establish a TPRM program and complete security assessments for all sub-processors",
            "Inventory and disclose all fourth parties with access to client data",
            "Include equivalent security obligations in all sub-processor contracts",
            "Confirm physical access controls and environmental safeguards for data facilities",
        ],
    },
    {
        "id": "N", "name": "AI & Emerging Technology Risk", "risk": "Critical",
        "critical": 9, "high": 4, "medium": 0, "total": 14,
        "gaps": [
            {"ref": "N.1.1", "text": "No formal AI usage policy approved by senior leadership", "tier": "Critical"},
            {"ref": "N.1.2", "text": "No AI risk assessment; roles and responsibilities for AI governance undefined", "tier": "Critical"},
            {"ref": "N.2.1", "text": "Client data may be used to train AI models without explicit consent", "tier": "Critical"},
            {"ref": "N.2.2", "text": "No technical controls preventing PII from entering AI models", "tier": "Critical"},
            {"ref": "N.2.3", "text": "AI outputs not validated before use in customer-affecting decisions", "tier": "Critical"},
            {"ref": "N.3.1", "text": "Third-party AI services lack DPA confirming no training on client data", "tier": "Critical"},
            {"ref": "N.4.1", "text": "AI systems lack RBAC and MFA; not hardened at deployment", "tier": "Critical"},
            {"ref": "N.4.2", "text": "No controls to detect prompt injection or adversarial input attacks", "tier": "Critical"},
            {"ref": "N.5.1", "text": "No tamper-protected audit logs for AI interactions", "tier": "Critical"},
        ],
        "recs": [
            "Publish an AI usage policy (acceptable use, prohibited inputs, output handling) approved by CISO/DPO",
            "Conduct a formal AI risk assessment; define AI governance roles and incident response procedures",
            "Require DPA from all third-party AI providers confirming client data is not used for model training",
            "Implement PII masking/tokenization before data reaches AI models; enable prompt/output logging",
            "Deploy prompt injection detection and output guardrails; enforce RBAC + MFA on all AI system access",
            "Maintain tamper-protected audit logs for all AI interactions with regular anomaly review",
        ],
    },
]


# ── Helpers ────────────────────────────────────────────────────────────────────
def normalize_response(val) -> str:
    if val is None:
        return "—"
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%m/%d/%Y")
    s = str(val).strip().lower()
    if s in YES_VALS:  return "Yes"
    if s in NO_VALS:   return "No"
    if s in NA_VALS:   return "N/A"
    if s in PART_VALS: return "Partial"
    if not s or s == "—": return "—"
    return str(val).strip()


def extract_domain(key) -> str:
    if not key:
        return ""
    s = str(key).strip()
    # "A — ORGANIZATIONAL MANAGEMENT" style headers
    if " — " in s:
        return s.split(" — ")[0].strip().upper()
    if s and s[0].isalpha():
        return s[0].upper()
    return ""


def compliance_score(items: list) -> int:
    scored = [i for i in items if i["norm"] not in ("—", "N/A")]
    if not scored:
        return 0
    earned = sum(100 if i["norm"] == "Yes" else (50 if i["norm"] == "Partial" else 0)
                 for i in scored)
    return round(earned / len(scored))


def risk_rating(score: int) -> tuple:
    for threshold, label, color in RATING_THRESHOLDS:
        if score >= threshold:
            return label, color
    return "🔴 Critical", "#A32D2D"


def pill(text: str, style: str) -> str:
    return f'<span class="response-pill" style="{style}">{text}</span>'


def resp_pill(norm: str) -> str:
    return pill(norm if norm != "—" else "—", RESP_PILL.get(norm, RESP_PILL["—"]))


def tier_pill(tier: str) -> str:
    return pill(tier, TIER_PILL.get(tier, "background:#F1EFE8;color:#888780"))


# ── Parsers ────────────────────────────────────────────────────────────────────
def parse_part1(wb) -> dict:
    """Parse Part 1 — Contact & Engagement information."""
    if "Part 1" not in wb.sheetnames:
        return {}
    ws = wb["Part 1"]
    rows = list(ws.iter_rows(values_only=True))
    meta = {"title": "", "sections": {}, "items": []}

    current_section = ""
    for row in rows:
        if not any(v is not None for v in row):
            continue
        key, question, response = row[0], row[1], row[2] if len(row) > 2 else None

        # Title row
        if isinstance(key, str) and "TPCRA" in str(key) and not meta["title"]:
            meta["title"] = str(key).strip()
            continue

        # Section header
        if isinstance(key, str) and key.startswith("SECTION"):
            current_section = str(key).strip()
            meta["sections"].setdefault(current_section, [])
            continue

        # Column header row — skip
        if key == "#":
            continue

        # Data row
        if question and str(question).strip():
            item = {
                "key":      str(key).strip() if key else "",
                "section":  current_section,
                "question": str(question).strip(),
                "response": str(response).strip() if response is not None else "",
                "other":    str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
                "tier":     str(row[4]).strip() if len(row) > 4 and row[4] and str(row[4]) != "—" else "",
            }
            meta["items"].append(item)
            if current_section:
                meta["sections"][current_section].append(item)

    return meta


def parse_part2(wb) -> dict:
    """Parse Part 2 — Security questionnaire with responses, tiers, and remarks."""
    if "Part 2" not in wb.sheetnames:
        return {}
    ws = wb["Part 2"]
    rows = list(ws.iter_rows(values_only=True))

    result = {"title": "", "domains": {}, "items": []}
    current_domain = ""
    current_sub = ""

    for row in rows:
        if not any(v is not None for v in row):
            continue
        key = row[0]
        question  = row[1] if len(row) > 1 else None
        response  = row[2] if len(row) > 2 else None
        other     = row[3] if len(row) > 3 else None
        tier      = row[4] if len(row) > 4 else None

        key_s = str(key).strip() if key else ""

        # Title row
        if isinstance(key, str) and "TPCRA" in key and not result["title"]:
            result["title"] = key.strip()
            continue

        # Column header row
        if key_s == "#":
            continue

        # Domain header: "A — ORGANIZATIONAL MANAGEMENT"
        if isinstance(key, str) and " — " in key and key[0].isalpha() and len(key.split(" — ")[0]) == 1:
            letter = key.split(" — ")[0].strip().upper()
            name   = key.split(" — ", 1)[1].strip()
            current_domain = letter
            current_sub = ""
            result["domains"].setdefault(letter, {
                "name": DOMAIN_MAP.get(letter, name),
                "items": []
            })
            continue

        # Sub-section label rows (no response, key has a dot or is descriptive text)
        if isinstance(key, str) and question and response is None and tier is None:
            current_sub = str(question).strip() if question else ""
            continue

        # Question row — must have a proper key like A.1, B.2.3, N.1.1 etc.
        if key_s and question and str(question).strip():
            # Determine domain from key
            domain_letter = key_s[0].upper() if key_s[0].isalpha() else current_domain

            tier_s = str(tier).strip() if tier and str(tier).strip() not in ("—", "None", "") else ""
            norm   = normalize_response(response)

            item = {
                "key":      key_s,
                "domain":   domain_letter,
                "domain_name": DOMAIN_MAP.get(domain_letter, domain_letter),
                "sub":      current_sub,
                "question": str(question).strip(),
                "response": str(response).strip() if response else "",
                "norm":     norm,
                "other":    str(other).strip() if other else "",
                "tier":     tier_s,
            }

            result["items"].append(item)
            if domain_letter in result["domains"]:
                result["domains"][domain_letter]["items"].append(item)
            elif domain_letter:
                result["domains"].setdefault(domain_letter, {
                    "name": DOMAIN_MAP.get(domain_letter, domain_letter),
                    "items": [item]
                })

    return result


def parse_evidence(wb) -> list:
    """Parse Evidence checklist sheet."""
    if "Evidence" not in wb.sheetnames:
        return []
    ws = wb["Evidence"]
    rows = list(ws.iter_rows(values_only=True))
    items = []
    for row in rows[2:]:
        if not any(v is not None for v in row):
            continue
        num, evidence, guidance, status, remarks, required_for = (
            row[0], row[1], row[2], row[3], row[4], row[5] if len(row) > 5 else None
        )
        if evidence:
            status_norm = "Submitted" if str(status or "").strip().lower() in EVIDENCE_STATUS else (
                str(status).strip() if status else "Pending"
            )
            items.append({
                "num":          str(num).strip() if num else "",
                "evidence":     str(evidence).strip(),
                "guidance":     str(guidance).strip() if guidance else "",
                "status":       status_norm,
                "remarks":      str(remarks).strip() if remarks else "",
                "required_for": str(required_for).strip() if required_for else "",
            })
    return items


def extract_contact(p1_items: list) -> dict:
    """Pull key contact fields from Part 1 items."""
    contact = {"vendor": "", "rep": "", "email": "", "engagement": ""}
    for item in p1_items:
        # Skip column header row
        if item.get("key") == "#":
            continue
        # Normalise question text — strip trailing asterisks and whitespace
        q = item["question"].lower().rstrip(" *")
        r = item["response"]
        # Skip empty or placeholder values (case-insensitive)
        if not r or r.strip().lower() in ("", "none", "—", "response", "n/a", "not applicable"):
            continue
        if "company name" in q:
            contact["vendor"] = r.strip()
        elif "authorized representative" in q and "email" not in q:
            contact["rep"] = r.strip()
        elif "email" in q and "representative" in q:
            contact["email"] = r.strip()
        elif "description of the engagement" in q:
            contact["engagement"] = (r[:120] + "…" if len(r) > 120 else r).strip()
    return contact


def make_sample_excel() -> bytes:
    """Generate a minimal sample Excel template matching TPCRA v3.0 format."""
    p2_rows = [
        ("TPCRA Questionnaire - Part 2  |  v3.0  |  Response options: Yes / No / Partial / N/A", None, None, None, None, None),
        ("#", "Statement / Question", "Response\n(Yes/No/Partial/N/A)", "Other Information\n(Remarks & Evidence)", "Risk\nTier", "Comments\nRequired"),
        ("A — ORGANIZATIONAL MANAGEMENT", None, None, None, None, None),
        ("A.1", "IT Security policies and procedures are formally established and documented.", None, None, "Critical", "—"),
        ("A.2", "IT Security policies and procedures are reviewed at least annually.", None, None, "High", "—"),
        ("A.5", "IT Security policies are formally acknowledged by all employees and contractors.", None, None, "High", "—"),
        ("A.6", "IT Security policies comply with relevant regulatory requirements.", None, None, "Critical", "—"),
        ("B — HUMAN RESOURCE MANAGEMENT", None, None, None, None, None),
        ("B.2", "IT security awareness training is provided to ALL employees.", None, None, "High", "—"),
        ("B.3", "New hires are subjected to background screening and must sign an NDA.", None, None, "Critical", "—"),
        ("B.5", "A formal employee offboarding process revokes all access on the last working day.", None, None, "Critical", "—"),
        ("C — INFRASTRUCTURE SECURITY", None, None, None, None, None),
        ("C.1.2", "Clear network segmentation exists between web, application, and database tiers.", None, None, "Critical", "—"),
        ("C.1.5", "All inter-system communication uses encrypted protocols only (HTTPS, SFTP, TLS 1.2+).", None, None, "Critical", "—"),
        ("C.5.1", "All remote connections require MFA and encrypted communications.", None, None, "Critical", "—"),
        ("C.6.2", "Security patches are applied within defined SLA timelines based on criticality.", None, None, "Critical", "—"),
        ("C.7.2", "Anti-malware solutions are deployed on all user computers, servers, and endpoints.", None, None, "Critical", "—"),
        ("D — DATA PROTECTION", None, None, None, None, None),
        ("D.1.3", "Technical measures isolate company data from other clients.", None, None, "Critical", "—"),
        ("D.1.10", "A DLP solution or equivalent is deployed to prevent unauthorized data exfiltration.", None, None, "High", "—"),
        ("D.2.1", "Strong cryptographic protocols protect sensitive data in transit.", None, None, "Critical", "—"),
        ("E — ACCESS MANAGEMENT", None, None, None, None, None),
        ("E.1.1", "Access to critical systems is granted only on a need-to-know basis.", None, None, "Critical", "—"),
        ("E.1.7", "Sharing of user IDs or privileged accounts is strictly prohibited.", None, None, "Critical", "—"),
        ("E.1.8", "MFA is enforced for all remote access and privileged account usage.", None, None, "Critical", "—"),
        ("J — INCIDENT RESPONSE", None, None, None, None, None),
        ("J.2", "A documented Incident Response Plan (IRP) is established.", None, None, "Critical", "—"),
        ("J.3", "The IRP is tested at least annually.", None, None, "Critical", "—"),
        ("N — AI & EMERGING TECHNOLOGY RISK", None, None, None, None, None),
        ("N.1.1", "A formal AI usage policy is in place covering acceptable use and accountability.", None, None, "Critical", "—"),
        ("N.2.1", "Data will NOT be used to train AI models without explicit written consent.", None, None, "Critical", "—"),
    ]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df = pd.DataFrame(p2_rows)
        df.to_excel(writer, index=False, header=False, sheet_name="Part 2")
    return buf.getvalue()


# ── PDF generation ────────────────────────────────────────────────────────────
def generate_vendor_pdf(
    live_gaps: list,
    vendor_name: str = "",
    assessment_date: str = "",
    total_questions: int = 0,
    gap_db: list = None,
) -> bytes:
    """
    Generate a vendor-facing gap report PDF from live questionnaire gap data.
    Page 1 — KPI summary + domain summary table.
    Page 2+ — Detailed gap listing per domain with high-level recommendations.
    gap_db — optional GAP_DB list; used to render per-domain recommendations.
    """
    from reportlab.platypus import PageBreak
    from collections import OrderedDict

    buf = io.BytesIO()
    PAGE_W, PAGE_H = A4
    MARGIN = 16 * mm
    usable_w = PAGE_W - 2 * MARGIN

    if not assessment_date:
        assessment_date = datetime.date.today().strftime("%d %B %Y")
    vendor_line = vendor_name.strip() if vendor_name and vendor_name.strip() else "Not provided (complete Section 1 of Part 1)"

    # ── colour palette ─────────────────────────────────────────────────────────
    C_RED    = colors.HexColor("#A32D2D")
    C_AMB    = colors.HexColor("#854F0B")
    C_BLU    = colors.HexColor("#185FA5")
    C_GRN    = colors.HexColor("#3B6D11")
    C_GRY    = colors.HexColor("#6c757d")
    C_BDR    = colors.HexColor("#e9ecef")
    C_HDR_BG = colors.HexColor("#1a1a2e")
    C_TBL_BG = colors.HexColor("#f8f9fa")
    C_REC_BG = colors.HexColor("#F0F4FF")
    C_REC_BD = colors.HexColor("#c7d2fe")

    TIER_FG = {"Critical": C_RED, "High": C_AMB, "Medium": C_BLU, "Low": C_GRN}
    TIER_BG = {
        "Critical": colors.HexColor("#FCEBEB"),
        "High":     colors.HexColor("#FAEEDA"),
        "Medium":   colors.HexColor("#E6F1FB"),
        "Low":      colors.HexColor("#EAF3DE"),
    }
    RESP_FG = {"No": C_RED, "Partial": C_AMB, "N/A": C_GRY, "—": C_GRY}
    RESP_BG = {
        "No":      colors.HexColor("#FCEBEB"),
        "Partial": colors.HexColor("#FAEEDA"),
        "N/A":     colors.HexColor("#F1EFE8"),
        "—":       colors.HexColor("#F1EFE8"),
    }

    def _s(name, **kw):
        s = ParagraphStyle(name)
        for k, v in kw.items():
            setattr(s, k, v)
        return s

    s_title  = _s("title",  fontSize=13, fontName="Helvetica-Bold",  textColor=colors.white,             alignment=TA_LEFT,  leading=17)
    s_body   = _s("body",   fontSize=8,  fontName="Helvetica",        textColor=colors.HexColor("#333"), leading=11)
    s_small  = _s("small",  fontSize=7,  fontName="Helvetica",        textColor=C_GRY,                   leading=10)
    s_ref    = _s("ref",    fontSize=7,  fontName="Helvetica-Bold",   textColor=C_GRY)
    s_q      = _s("q",      fontSize=7.5,fontName="Helvetica",        textColor=colors.HexColor("#222"), leading=10)
    s_rmk    = _s("rmk",    fontSize=6.5,fontName="Helvetica-Oblique",textColor=C_GRY,                   leading=9)
    s_sec    = _s("sec",    fontSize=7.5,fontName="Helvetica-Bold",   textColor=C_GRY,                   spaceBefore=5, spaceAfter=2)
    s_note   = _s("note",   fontSize=7.5,fontName="Helvetica-Oblique",textColor=C_GRY,                   leading=11)
    s_footer = _s("footer", fontSize=7,  fontName="Helvetica",        textColor=C_GRY,                   alignment=TA_CENTER)
    s_rec_hd = _s("rec_hd", fontSize=7.5,fontName="Helvetica-Bold",   textColor=colors.HexColor("#1a1a2e"), leading=10)
    s_rec    = _s("rec",    fontSize=7,  fontName="Helvetica",        textColor=colors.HexColor("#333"),  leading=10)

    # Build a quick lookup: domain letter → list of recommendation strings
    rec_lookup: dict = {}
    if gap_db:
        for entry in gap_db:
            if entry.get("recs"):
                rec_lookup[entry["id"]] = entry["recs"]

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=14*mm,
    )
    story = []

    # ── shared helpers ─────────────────────────────────────────────────────────
    def _footer_row():
        ft = Table([[
            Paragraph("TPCRA v3.0  |  For vendor assessment use only  |  Confidential", s_footer),
            Paragraph(f"Generated: {assessment_date}",
                      _s("fr", fontSize=7, fontName="Helvetica", textColor=C_GRY, alignment=TA_RIGHT)),
        ]], colWidths=[usable_w * 0.72, usable_w * 0.28])
        ft.setStyle(TableStyle([
            ("TOPPADDING", (0,0),(-1,-1), 5),
            ("LINEABOVE",  (0,0),(-1,-1), 0.5, C_BDR),
        ]))
        return ft

    def _header_banner():
        t = Table([[Paragraph(
            "Third-Party Cyber Risk Assessment  —  Gap Assessment Report", s_title
        )]], colWidths=[usable_w])
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0),(-1,-1), C_HDR_BG),
            ("TOPPADDING",   (0,0),(-1,-1), 10),
            ("BOTTOMPADDING",(0,0),(-1,-1), 10),
            ("LEFTPADDING",  (0,0),(-1,-1), 12),
            ("RIGHTPADDING", (0,0),(-1,-1), 12),
        ]))
        return t

    def _meta_row():
        n_gaps  = len(live_gaps)
        pct_gap = f"{round(n_gaps / total_questions * 100)}%" if total_questions else "—"
        t = Table([[
            Paragraph(f"<b>Vendor:</b> {vendor_line}", s_body),
            Paragraph(f"<b>Assessment date:</b> {assessment_date}", s_body),
            Paragraph(f"<b>Total gaps:</b> {n_gaps} of {total_questions} ({pct_gap})", s_body),
            Paragraph("<b>Framework:</b> TPCRA v3.0", s_body),
        ]], colWidths=[usable_w*0.30, usable_w*0.26, usable_w*0.26, usable_w*0.18])
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0),(-1,-1), C_TBL_BG),
            ("BOX",          (0,0),(-1,-1), 0.5, C_BDR),
            ("TOPPADDING",   (0,0),(-1,-1), 6),
            ("BOTTOMPADDING",(0,0),(-1,-1), 6),
            ("LEFTPADDING",  (0,0),(-1,-1), 8),
            ("RIGHTPADDING", (0,0),(-1,-1), 8),
            ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ]))
        return t

    # ── aggregate counts ───────────────────────────────────────────────────────
    def _cnt(items, tier=None, resp=None):
        return sum(1 for i in items
                   if (tier is None or i["tier"] == tier)
                   and (resp is None or i["norm"] == resp))

    n_part  = _cnt(live_gaps, resp="Partial")
    n_na    = _cnt(live_gaps, resp="N/A")
    n_unans = _cnt(live_gaps, resp="—")

    # Group by domain preserving DOMAIN_MAP order
    by_domain = OrderedDict()
    for letter in DOMAIN_MAP.keys():
        items = [i for i in live_gaps if i["domain"] == letter]
        if items:
            by_domain[letter] = {"name": DOMAIN_MAP.get(letter, letter), "items": items}

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 1 — KPI summary + domain breakdown table
    # ══════════════════════════════════════════════════════════════════════════
    story.append(_header_banner())
    story.append(Spacer(1, 2.5*mm))
    story.append(_meta_row())
    story.append(Spacer(1, 3*mm))

    # KPI strip — response type counts only
    def _kpi(val, label, color="#333333"):
        return Paragraph(
            f'<font size="14" color="{color}"><b>{val}</b></font><br/>'
            f'<font size="7" color="#6c757d">{label}</font>',
            _s(f"kpi_{label[:5]}", fontName="Helvetica", alignment=TA_CENTER, leading=18),
        )

    kpi_tbl = Table([[
        _kpi(len(live_gaps), "Total gaps"),
        _kpi(n_part,  "Partial",    "#854F0B"),
        _kpi(n_na,    "N/A",        "#5F5E5A"),
        _kpi(n_unans, "Unanswered", "#888780"),
    ]], colWidths=[usable_w / 4] * 4)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,-1), colors.white),
        ("BOX",          (0,0),(-1,-1), 0.5, C_BDR),
        ("INNERGRID",    (0,0),(-1,-1), 0.5, C_BDR),
        ("TOPPADDING",   (0,0),(-1,-1), 7),
        ("BOTTOMPADDING",(0,0),(-1,-1), 7),
        ("ALIGN",        (0,0),(-1,-1), "CENTER"),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 3.5*mm))

    # Domain summary table
    story.append(Paragraph("GAP SUMMARY BY DOMAIN", s_sec))

    col_w_sum = [
        usable_w*0.05,  # ID
        usable_w*0.27,  # Name
        usable_w*0.08,  # Total
        usable_w*0.08,  # Partial
        usable_w*0.08,  # N/A
        usable_w*0.08,  # Unanswered
        usable_w*0.36,  # Top gap preview
    ]
    sum_rows = [[
        Paragraph("#",           s_small),
        Paragraph("Domain",      s_small),
        Paragraph("Total",       s_small),
        Paragraph("Partial",     s_small),
        Paragraph("N/A",         s_small),
        Paragraph("Unanswered",  s_small),
        Paragraph("Top gap",     s_small),
    ]]

    for letter, dom in by_domain.items():
        di = dom["items"]
        dtotal = len(di)
        dp  = _cnt(di, resp="Partial")
        dna = _cnt(di, resp="N/A")
        dun = _cnt(di, resp="—")
        top = next((i for i in di if i["tier"] in ("Critical","High")), di[0])
        top_text = top["question"][:80] + ("..." if len(top["question"]) > 80 else "")

        def _num(n, fg=C_GRY):
            return Paragraph(str(n) if n else "—",
                _s(f"n{letter}{n}", fontSize=7.5,
                   fontName="Helvetica-Bold" if n else "Helvetica",
                   textColor=fg if n else C_GRY, alignment=TA_CENTER))

        sum_rows.append([
            Paragraph(letter, s_ref),
            Paragraph(dom["name"], s_body),
            _num(dtotal, C_GRY),
            _num(dp,     C_AMB),
            _num(dna,    C_GRY),
            _num(dun,    C_GRY),
            Paragraph(top_text, _s(f"tg{letter}", fontSize=6.5, fontName="Helvetica",
                                   textColor=colors.HexColor("#444"), leading=9)),
        ])

    sum_tbl = Table(sum_rows, colWidths=col_w_sum, repeatRows=1)
    sum_ts = [
        ("BACKGROUND",   (0,0),(-1,0),  C_TBL_BG),
        ("BOX",          (0,0),(-1,-1), 0.5, C_BDR),
        ("INNERGRID",    (0,0),(-1,-1), 0.3, C_BDR),
        ("TOPPADDING",   (0,0),(-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1), 3),
        ("LEFTPADDING",  (0,0),(-1,-1), 4),
        ("RIGHTPADDING", (0,0),(-1,-1), 4),
        ("VALIGN",       (0,0),(-1,-1), "TOP"),
        ("ALIGN",        (2,1),(5,-1),  "CENTER"),
    ]
    for r in range(1, len(sum_rows)):
        if r % 2 == 0:
            sum_ts.append(("BACKGROUND", (0,r),(-1,r), colors.HexColor("#fafafa")))
    sum_tbl.setStyle(TableStyle(sum_ts))
    story.append(sum_tbl)
    story.append(Spacer(1, 3*mm))

    # Expectations note
    story.append(Table([[Paragraph(
        "<b>Vendor expectations:</b> All Critical-tier gaps must be remediated or a documented compensating "
        "control provided before onboarding can proceed. High-tier gaps require a remediation plan with "
        "committed timelines within 30 days. Evidence must be submitted per the TPCRA v3.0 Evidence Checklist. "
        "This report is confidential and used solely for third-party risk assessment purposes.",
        s_note,
    )]], colWidths=[usable_w], style=TableStyle([
        ("BACKGROUND",   (0,0),(-1,-1), C_REC_BG),
        ("BOX",          (0,0),(-1,-1), 0.5, C_REC_BD),
        ("TOPPADDING",   (0,0),(-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("LEFTPADDING",  (0,0),(-1,-1), 9),
        ("RIGHTPADDING", (0,0),(-1,-1), 9),
    ])))
    story.append(Spacer(1, 3*mm))
    story.append(_footer_row())

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 2+ — Detailed gap listing per domain
    # ══════════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(_header_banner())
    story.append(Spacer(1, 2.5*mm))
    story.append(_meta_row())
    story.append(Spacer(1, 3.5*mm))
    story.append(Paragraph("DETAILED GAP LISTING BY DOMAIN", s_sec))
    story.append(Spacer(1, 1.5*mm))

    col_w_det = [usable_w*0.07, usable_w*0.09, usable_w*0.09, usable_w*0.75]

    for letter, dom in by_domain.items():
        ditems_sorted = sorted(dom["items"], key=lambda x: (
            {"Critical":0,"High":1,"Medium":2,"Low":3,"":4}.get(x["tier"], 4), x["key"]
        ))
        dc = _cnt(dom["items"], tier="Critical")
        dh = _cnt(dom["items"], tier="High")
        dm2 = _cnt(dom["items"], tier="Medium")
        dl = _cnt(dom["items"], tier="Low")

        # Domain section header
        dom_hdr = Table([[Paragraph(
            f"{letter}  —  {dom['name']}     "
            f"Critical: {dc}   High: {dh}   Medium: {dm2}   Low: {dl}   "
            f"Total: {len(dom['items'])}",
            _s(f"dh{letter}", fontSize=8, fontName="Helvetica-Bold",
               textColor=colors.white, leading=11),
        )]], colWidths=[usable_w])
        dom_hdr.setStyle(TableStyle([
            ("BACKGROUND",   (0,0),(-1,-1), colors.HexColor("#2d3250")),
            ("TOPPADDING",   (0,0),(-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
            ("LEFTPADDING",  (0,0),(-1,-1), 9),
            ("RIGHTPADDING", (0,0),(-1,-1), 9),
        ]))
        story.append(dom_hdr)

        det_rows = [[
            Paragraph("Ref",      s_small),
            Paragraph("Tier",     s_small),
            Paragraph("Response", s_small),
            Paragraph("Control / Question", s_small),
        ]]
        for item in ditems_sorted:
            tier_fg = TIER_FG.get(item["tier"], C_GRY)
            tier_bg = TIER_BG.get(item["tier"], colors.HexColor("#F1EFE8"))
            resp_fg = RESP_FG.get(item["norm"], C_GRY)
            resp_bg = RESP_BG.get(item["norm"], colors.HexColor("#F1EFE8"))
            resp_label = item["norm"] if item["norm"] != "—" else "Unanswered"

            q_parts = [Paragraph(item["question"], s_q)]
            if item.get("other"):
                q_parts.append(Paragraph(f"Remarks: {item['other']}", s_rmk))
            q_cell = Table([[p] for p in q_parts], colWidths=[col_w_det[3]])
            q_cell.setStyle(TableStyle([
                ("TOPPADDING",   (0,0),(-1,-1), 1),
                ("BOTTOMPADDING",(0,0),(-1,-1), 1),
                ("LEFTPADDING",  (0,0),(-1,-1), 0),
                ("RIGHTPADDING", (0,0),(-1,-1), 0),
            ]))
            det_rows.append([
                Paragraph(item["key"],
                          _s(f"ref{item['key']}", fontSize=7, fontName="Helvetica-Bold", textColor=C_GRY)),
                Paragraph(item["tier"] if item["tier"] else "—",
                          _s(f"tf{item['key']}", fontSize=7, fontName="Helvetica-Bold",
                             textColor=tier_fg, alignment=TA_CENTER)),
                Paragraph(resp_label,
                          _s(f"rf{item['key']}", fontSize=7, fontName="Helvetica-Bold",
                             textColor=resp_fg, alignment=TA_CENTER)),
                q_cell,
            ])

        det_tbl = Table(det_rows, colWidths=col_w_det, repeatRows=1)
        det_ts = [
            ("BACKGROUND",   (0,0),(-1,0),  C_TBL_BG),
            ("BOX",          (0,0),(-1,-1), 0.5, C_BDR),
            ("INNERGRID",    (0,0),(-1,-1), 0.3, C_BDR),
            ("TOPPADDING",   (0,0),(-1,-1), 4),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
            ("LEFTPADDING",  (0,0),(-1,-1), 4),
            ("RIGHTPADDING", (0,0),(-1,-1), 4),
            ("VALIGN",       (0,0),(-1,-1), "TOP"),
            ("ALIGN",        (1,1),(2,-1),  "CENTER"),
        ]
        for r_idx, item in enumerate(ditems_sorted, 1):
            det_ts.append(("BACKGROUND", (1,r_idx),(1,r_idx), TIER_BG.get(item["tier"], colors.HexColor("#F1EFE8"))))
            det_ts.append(("BACKGROUND", (2,r_idx),(2,r_idx), RESP_BG.get(item["norm"], colors.HexColor("#F1EFE8"))))
            if r_idx % 2 == 0:
                det_ts.append(("BACKGROUND", (0,r_idx),(0,r_idx), colors.HexColor("#fafafa")))
                det_ts.append(("BACKGROUND", (3,r_idx),(3,r_idx), colors.HexColor("#fafafa")))
        det_tbl.setStyle(TableStyle(det_ts))
        story.append(det_tbl)

        # ── Per-domain recommendations from GAP_DB ─────────────────────────────
        domain_recs = rec_lookup.get(letter, [])
        if domain_recs:
            rec_hdr_tbl = Table(
                [[Paragraph("Recommendations", s_rec_hd)]],
                colWidths=[usable_w],
            )
            rec_hdr_tbl.setStyle(TableStyle([
                ("BACKGROUND",   (0,0),(-1,-1), C_REC_BG),
                ("LEFTPADDING",  (0,0),(-1,-1), 9),
                ("RIGHTPADDING", (0,0),(-1,-1), 9),
                ("TOPPADDING",   (0,0),(-1,-1), 4),
                ("BOTTOMPADDING",(0,0),(-1,-1), 4),
                ("LINEABOVE",    (0,0),(-1,-1), 0.5, C_REC_BD),
                ("LINEBELOW",    (0,0),(-1,-1), 0.5, C_REC_BD),
            ]))
            story.append(rec_hdr_tbl)

            rec_rows = []
            for j, rec_text in enumerate(domain_recs, 1):
                bullet = Paragraph(str(j), _s(f"bn{letter}{j}", fontSize=7,
                                               fontName="Helvetica-Bold",
                                               textColor=colors.HexColor("#1a1a2e"),
                                               alignment=TA_CENTER))
                text   = Paragraph(rec_text, s_rec)
                rec_rows.append([bullet, text])

            rec_tbl = Table(rec_rows, colWidths=[usable_w*0.04, usable_w*0.96])
            rec_tbl.setStyle(TableStyle([
                ("BACKGROUND",   (0,0),(-1,-1), C_REC_BG),
                ("BOX",          (0,0),(-1,-1), 0.5, C_REC_BD),
                ("LINEBELOW",    (0,0),(-1,-1), 0.3, C_BDR),
                ("TOPPADDING",   (0,0),(-1,-1), 3),
                ("BOTTOMPADDING",(0,0),(-1,-1), 3),
                ("LEFTPADDING",  (0,0),(-1,-1), 4),
                ("RIGHTPADDING", (0,0),(-1,-1), 6),
                ("VALIGN",       (0,0),(-1,-1), "TOP"),
            ]))
            story.append(rec_tbl)

        story.append(Spacer(1, 4*mm))

    story.append(_footer_row())
    doc.build(story)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.header("Upload questionnaire")
    uploaded = st.file_uploader(
        "Use TPCRA v3.0 questionnaire",
        type=["xlsx", "xls"],
        help="Upload a completed TPCRA v3.0 questionnaire."
    )
    st.divider()
    st.download_button(
        "⬇ Download sample template",
        data=make_sample_excel(),
        file_name="TPCRA_v3.0_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Download a minimal TPCRA v3.0 questionnaire template to fill in.",
        use_container_width=True,
    )

# ── Empty state ────────────────────────────────────────────────────────────────
if not uploaded:
    st.title("🔐 Third Party Risk Assessment Dashboard")
    st.markdown("Upload a completed TPCRA v3.0 questionnaire from the sidebar to generate the dashboard.")
    st.divider()
    c1, c2, c3 = st.columns(3)
    c1.info("**Overview**\nCompliance scores, response distribution, domain charts, and per-domain question drill-down.")
    c2.info("**Gap List**\nAll No / N/A / Partial / unanswered controls grouped by domain, with recommendations and PDF export.")
    c3.info("**Evidence Checklist**\nEvidence submission status against the 14 required evidence items from the TPCRA v3.0 checklist.")
    c4, c5, c6 = st.columns(3)
    c4.info("**Engagement Info**\nVendor contact details, engagement description, data handling, and transmission methods from Part 1.")
    c5.markdown("")
    c6.markdown("")
    st.stop()

# ── Load workbook ──────────────────────────────────────────────────────────────
try:
    wb = load_workbook(uploaded, read_only=True, data_only=True)
except Exception as e:
    st.error(f"Could not open file: {e}")
    st.stop()

p1_data  = parse_part1(wb)
p2_data  = parse_part2(wb)
evidence = parse_evidence(wb)

if not p2_data or not p2_data.get("items"):
    st.error("No Part 2 question data found. Ensure the file has a 'Part 2' sheet matching the TPCRA v3.0 format.")
    st.stop()

contact  = extract_contact(p1_data.get("items", []))
p2_items = p2_data["items"]
domains  = p2_data["domains"]

# ── Overall score (used in header and Overview tab) ────────────────────────────
overall_score = compliance_score(p2_items)
overall_rating_label, overall_rating_color = risk_rating(overall_score)

# ── Header ─────────────────────────────────────────────────────────────────────
h1, h2 = st.columns([4, 1])
with h1:
    vendor_label = contact["vendor"] or "Vendor"
    st.title(f"🔐 {vendor_label}")
    if contact["rep"] or contact["email"]:
        st.caption(f"{contact['rep']}  ·  {contact['email']}")
    if contact["engagement"]:
        st.caption(f"Engagement: {contact['engagement']}")
with h2:
    st.caption(p2_data.get("title", "TPCRA v3.0"))
    st.markdown(
        f'<div style="text-align:right;margin-top:4px">'
        f'<div style="font-size:11px;color:#6c757d;font-weight:600;'
        f'text-transform:uppercase;letter-spacing:0.05em;margin-bottom:2px">Compliance score</div>'
        f'<div style="font-size:28px;font-weight:700;color:{overall_rating_color};line-height:1.1">'
        f'{overall_score}%</div>'
        f'<div style="font-size:12px;color:{overall_rating_color};font-weight:600">'
        f'{overall_rating_label}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

# ── Response counts (used by Overview tab charts) ──────────────────────────────
n_yes  = sum(1 for i in p2_items if i["norm"] == "Yes")
n_no   = sum(1 for i in p2_items if i["norm"] == "No")
n_part = sum(1 for i in p2_items if i["norm"] == "Partial")
n_na   = sum(1 for i in p2_items if i["norm"] == "N/A")

# ── Tabs ───────────────────────────────────────────────────────────────────────
tab_part1, tab_overview, tab_gap_summary, tab_evidence = st.tabs([
    "Engagement Info", "Overview", "Gap List", "Evidence Checklist"
])

# ══════════════════════════════════════════════
# TAB 1 — OVERVIEW  (includes By Domain drill-down)
# ══════════════════════════════════════════════
with tab_overview:
    col_bar, col_right = st.columns([3, 2])

    with col_bar:
        st.subheader("Compliance by domain")
        dom_rows = []
        for letter, dom in domains.items():
            items = dom["items"]
            if not items:
                continue
            sc = compliance_score(items)
            rl, rc = risk_rating(sc)
            dom_rows.append({
                "Domain":  dom["name"],
                "Yes":     sum(1 for i in items if i["norm"] == "Yes"),
                "Partial": sum(1 for i in items if i["norm"] == "Partial"),
                "No":      sum(1 for i in items if i["norm"] == "No"),
                "N/A":     sum(1 for i in items if i["norm"] == "N/A"),
                "Score":   sc,
                "Rating":  rl,
            })
        dom_df = pd.DataFrame(dom_rows)

        if not dom_df.empty:
            melt = dom_df.melt(
                id_vars="Domain", value_vars=["Yes", "Partial", "No", "N/A"],
                var_name="Response", value_name="Count"
            ).query("Count > 0")
            fig_bar = px.bar(
                melt, x="Count", y="Domain", color="Response",
                orientation="h",
                color_discrete_map=RESP_COLORS,
                category_orders={"Response": ["Yes", "Partial", "No", "N/A"]},
                labels={"Count": "Questions", "Domain": ""},
                height=max(380, len(dom_df) * 42),
            )
            fig_bar.update_layout(
                margin=dict(l=0, r=0, t=10, b=0),
                legend_title_text="Response",
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                yaxis={"categoryorder": "total ascending"},
                font=dict(size=12),
            )
            fig_bar.update_xaxes(showgrid=True, gridcolor="rgba(0,0,0,0.07)")
            st.plotly_chart(fig_bar, use_container_width=True)

    with col_right:
        st.subheader("Distribution")
        labels = [k for k, v in [("Yes",n_yes),("Partial",n_part),("No",n_no),("N/A",n_na)] if v > 0]
        values = [v for k, v in [("Yes",n_yes),("Partial",n_part),("No",n_no),("N/A",n_na)] if v > 0]
        fig_donut = go.Figure(go.Pie(
            labels=labels, values=values, hole=0.62,
            marker_colors=[RESP_COLORS[l] for l in labels],
            textinfo="label+percent",
            hovertemplate="%{label}: %{value}<extra></extra>",
        ))
        fig_donut.update_layout(
            margin=dict(l=0, r=0, t=10, b=0), showlegend=False,
            height=260, paper_bgcolor="rgba(0,0,0,0)", font=dict(size=12),
        )
        st.plotly_chart(fig_donut, use_container_width=True)

        st.subheader("Domain scores")
        if not dom_df.empty:
            for _, row in dom_df.sort_values("Score").iterrows():
                _, rc = risk_rating(row["Score"])
                st.markdown(
                    f'<div style="margin-bottom:7px">'
                    f'<div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:3px">'
                    f'<span style="color:#555">{row["Domain"]}</span>'
                    f'<span style="font-weight:600;color:{rc}">{row["Score"]}%</span></div>'
                    f'<div style="background:#f0f0f0;border-radius:4px;height:5px;overflow:hidden">'
                    f'<div style="width:{row["Score"]}%;height:100%;background:{rc};border-radius:4px"></div>'
                    f'</div></div>', unsafe_allow_html=True,
                )

    # ── By domain drill-down ────────────────────────────────────────────────────
    st.divider()
    st.subheader("By domain")

    domain_choices = [f"{l} — {domains[l]['name']}" for l in domains if domains[l]["items"]]
    if not domain_choices:
        st.info("No domain data found.")
    else:
        chosen = st.selectbox("Select domain", domain_choices)
        chosen_letter = chosen.split(" — ")[0]
        dom = domains[chosen_letter]
        items = dom["items"]

        n_y = sum(1 for i in items if i["norm"] == "Yes")
        n_n = sum(1 for i in items if i["norm"] == "No")
        n_p = sum(1 for i in items if i["norm"] == "Partial")
        n_a = sum(1 for i in items if i["norm"] == "N/A")

        dm1, dm2, dm3, dm4, dm5 = st.columns(5)
        dm1.metric("Questions",   len(items))
        dm2.metric("✅ Yes",      n_y)
        dm3.metric("❌ No",       n_n)
        dm4.metric("⚠️ Partial",  n_p)
        dm5.metric("➖ N/A",      n_a)

        st.divider()

        # Filter
        resp_f = st.multiselect(
            "Filter by response", ["Yes","No","Partial","N/A","—"],
            default=["Yes","No","Partial","N/A","—"], key="dom_resp_filter"
        )
        shown = [i for i in items if i["norm"] in resp_f]
        st.caption(f"Showing {len(shown)} of {len(items)} questions")

        current_sub = ""
        for item in shown:
            # Sub-section label
            if item["sub"] and item["sub"] != current_sub:
                current_sub = item["sub"]
                st.markdown(
                    f'<div style="margin:16px 0 6px;font-size:11px;font-weight:700;'
                    f'text-transform:uppercase;letter-spacing:0.06em;color:#6c757d">'
                    f'{current_sub}</div>', unsafe_allow_html=True
                )

            norm = item["norm"]
            bg = {"No": "#fff5f5", "Partial": "#fffaf5"}.get(norm, "transparent")
            border = {"No": "#f7c1c1", "Partial": "#FAC775"}.get(norm, "#e9ecef")
            tier_badge = tier_pill(item["tier"]) if item["tier"] else ""
            resp_badge = resp_pill(norm)
            key_line = item["key"]
            tier_part = "&nbsp;&nbsp;" + tier_badge if tier_badge else ""
            remarks_html = (
                '<div style="font-size:12px;color:#555;margin-top:6px;'
                'padding-top:6px;border-top:0.5px solid #e9ecef">'
                + item["other"] + "</div>"
            ) if item["other"] else ""
            card_html = (
                '<div style="padding:10px 14px;border-radius:8px;margin-bottom:6px;'
                'background:' + bg + ';border:0.5px solid ' + border + '">'
                '<div style="display:flex;justify-content:space-between;align-items:flex-start;gap:10px">'
                '<div style="flex:1">'
                '<div style="font-size:11px;color:#aaa;margin-bottom:3px">'
                + key_line + tier_part +
                '</div>'
                '<div style="font-size:13px;color:#333;line-height:1.55">' + item["question"] + '</div>'
                + remarks_html +
                '</div>'
                '<div style="flex-shrink:0;padding-top:2px">' + resp_badge + '</div>'
                '</div>'
                '</div>'
            )
            st.markdown(card_html, unsafe_allow_html=True)

            # Long free-text responses
            if item["response"] and len(item["response"]) > 50 and norm not in ("Yes","No","Partial","N/A","—"):
                with st.expander(f"View full response — {item['key']}"):
                    st.markdown(
                        f'<div style="font-size:13px;color:#444;line-height:1.6;white-space:pre-wrap">'
                        f'{item["response"]}</div>', unsafe_allow_html=True
                    )
# ══════════════════════════
# TAB 3 — EVIDENCE
# ══════════════════════════
with tab_evidence:
    if not evidence:
        st.info("No Evidence sheet found in the uploaded file.")
    else:
        submitted = sum(1 for e in evidence if e["status"] == "Submitted")
        pending   = len(evidence) - submitted

        ev1, ev2, ev3 = st.columns(3)
        ev1.metric("Total evidence items", len(evidence))
        ev2.metric("✅ Submitted",          submitted)
        ev3.metric("⏳ Pending",            pending)
        st.divider()

        rows_html = ""
        for ev in evidence:
            s = ev["status"]
            s_style = "background:#EAF3DE;color:#27500A" if s == "Submitted" else "background:#FAEEDA;color:#633806"
            rows_html += (
                "<tr>"
                '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:12px;color:#aaa;width:4%;vertical-align:top">' + str(ev['num']) + "</td>"
                '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:13px;color:#333;width:26%;vertical-align:top;font-weight:500">' + str(ev['evidence']) + "</td>"
                '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:12px;color:#555;width:35%;vertical-align:top;line-height:1.5">' + str(ev['guidance']) + "</td>"
                '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:12px;width:20%;vertical-align:top">' + str(ev['required_for']) + "</td>"
                '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;width:15%;vertical-align:top">'
                '<span class="response-pill" style="' + s_style + '">' + s + "</span></td>"
                "</tr>"
            )

        TH = "padding:9px 10px;text-align:left;font-size:11px;font-weight:600;color:#6c757d;text-transform:uppercase;letter-spacing:0.05em;border-bottom:1px solid #e9ecef"
        ev_table = (
            '<div style="border:1px solid #e9ecef;border-radius:10px;overflow:hidden">'
            '<table style="width:100%;border-collapse:collapse;table-layout:fixed">'
            '<thead><tr style="background:#f8f9fa">'
            '<th style="' + TH + ';width:4%">#</th>'
            '<th style="' + TH + ';width:26%">Evidence required</th>'
            '<th style="' + TH + ';width:35%">Guidance</th>'
            '<th style="' + TH + ';width:20%">Required for</th>'
            '<th style="' + TH + ';width:15%">Status</th>'
            "</tr></thead>"
            "<tbody>" + rows_html + "</tbody>"
            "</table></div>"
        )
        st.markdown(ev_table, unsafe_allow_html=True)


# ══════════════════════════
# TAB 4 — ENGAGEMENT INFO
# ══════════════════════════
with tab_part1:
    if not p1_data or not p1_data.get("items"):
        st.info("No Part 1 data found in the uploaded file. Ensure the file contains a 'Part 1' sheet.")
    else:
        p1_sections = p1_data.get("sections", {})
        if not p1_sections:
            sections_to_show = {"All questions": p1_data["items"]}
        else:
            sections_to_show = p1_sections

        HIDE_TIER_SECTIONS = {"SECTION 1 — CONTACT PERSON", "SECTION 2 — ENGAGEMENT INFORMATION"}

        for sec_name, sec_items in sections_to_show.items():
            if not sec_items:
                continue
            st.subheader(sec_name.replace("SECTION ", "").replace(" — ", " — ").title()
                         if "SECTION" in sec_name else sec_name)

            hide_tier = sec_name in HIDE_TIER_SECTIONS

            rows_html = ""
            for item in sec_items:
                r = item["response"] or "—"
                resp_preview = r if len(r) <= 100 else r[:97] + "…"
                tier_cell = (
                    ""
                    if hide_tier
                    else '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;width:10%;vertical-align:top">'
                         + (tier_pill(item["tier"]) if item.get("tier") else "")
                         + "</td>"
                )
                q_width  = "55%" if hide_tier else "48%"
                resp_width = "38%" if hide_tier else "35%"
                rows_html += (
                    "<tr>"
                    '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:12px;color:#aaa;width:7%;vertical-align:top">' + item['key'] + "</td>"
                    f'<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:13px;color:#333;width:{q_width};vertical-align:top;line-height:1.5">' + item['question'] + "</td>"
                    f'<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:13px;color:#555;width:{resp_width};vertical-align:top;line-height:1.5">' + resp_preview + "</td>"
                    + tier_cell +
                    "</tr>"
                )

            TH2 = "padding:9px 10px;text-align:left;font-size:11px;font-weight:600;color:#6c757d;text-transform:uppercase;letter-spacing:0.05em;border-bottom:1px solid #e9ecef"
            tier_header = "" if hide_tier else f'<th style="{TH2};width:10%">Tier</th>'
            q_width_h   = "55%" if hide_tier else "48%"
            resp_width_h = "38%" if hide_tier else "35%"
            p1_table = (
                '<div style="border:1px solid #e9ecef;border-radius:10px;overflow:hidden;margin-bottom:1.5rem">'
                '<table style="width:100%;border-collapse:collapse;table-layout:fixed">'
                '<thead><tr style="background:#f8f9fa">'
                f'<th style="{TH2};width:7%">#</th>'
                f'<th style="{TH2};width:{q_width_h}">Question</th>'
                f'<th style="{TH2};width:{resp_width_h}">Response</th>'
                + tier_header +
                "</tr></thead>"
                "<tbody>" + rows_html + "</tbody>"
                "</table></div>"
            )
            st.markdown(p1_table, unsafe_allow_html=True)


# ══════════════════════════
# TAB 5 — GAP SUMMARY
# ══════════════════════════
with tab_gap_summary:

    # ── Response style helpers (reuse existing pill constants) ──────────────────
    GS_RISK_BADGE = {
        "Critical": "background:#FCEBEB;color:#791F1F;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:600",
        "High":     "background:#FAEEDA;color:#633806;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:600",
        "Medium":   "background:#E6F1FB;color:#0C447C;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:600",
        "Low":      "background:#EAF3DE;color:#27500A;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:600",
    }
    GS_RESP_BADGE = {
        "No":      "background:#FCEBEB;color:#A32D2D;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:600",
        "N/A":     "background:#F1EFE8;color:#5F5E5A;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:600",
        "Partial": "background:#FAEEDA;color:#633806;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:600",
        "—":       "background:#F1EFE8;color:#888780;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:600",
    }
    TIER_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3, "": 4}
    TIER_BORDER = {
        "Critical": "#f7c1c1",
        "High":     "#FAC775",
        "Medium":   "#B5D4F4",
        "Low":      "#C0DD97",
        "":         "#e9ecef",
    }
    TIER_BG = {
        "Critical": "#fff5f5",
        "High":     "#fffaf5",
        "Medium":   "#f5f9ff",
        "Low":      "#f6fbee",
        "":         "#fafafa",
    }

    # ── Build live gap dataset from uploaded questionnaire ──────────────────────
    # Gaps = items where norm is No, N/A, or Partial (and unanswered "—")
    GAP_RESPONSES = {"No", "N/A", "Partial", "—"}

    live_gaps = [i for i in p2_items if i["norm"] in GAP_RESPONSES]

    # Counts by tier × response for KPIs
    def _count(items, tier=None, resp=None):
        return sum(
            1 for i in items
            if (tier is None or i["tier"] == tier)
            and (resp is None or i["norm"] == resp)
        )

    total_gaps    = len(live_gaps)

    st.subheader("Gap Summary Report")
    st.caption(
        "Controls with No, N/A, or Partial responses only — sourced from the uploaded questionnaire. "
        "Filter by tier or response to prioritise remediation."
    )

    # ── Counts from full question set ───────────────────────────────────────────
    total_questions_gs = len(p2_items)
    n_yes_gs    = sum(1 for i in p2_items if i["norm"] == "Yes")
    n_no_total  = sum(1 for i in p2_items if i["norm"] == "No")
    n_part_total = sum(1 for i in p2_items if i["norm"] == "Partial")
    n_na_total  = sum(1 for i in p2_items if i["norm"] == "N/A")
    n_unans_total = sum(1 for i in p2_items if i["norm"] == "—")

    # ── KPI strip ────────────────────────────────────────────────────────────────
    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Total questions", total_questions_gs)
    k2.metric("✅ Yes",          n_yes_gs)
    k3.metric("❌ No",           n_no_total)
    k4.metric("⚠️ Partial",      n_part_total)
    k5.metric("➖ N/A",          n_na_total)
    k6.metric("⬜ Unanswered",   n_unans_total)

    st.divider()

    # ── Filters ──────────────────────────────────────────────────────────────────
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        gs_tier_f = st.multiselect(
            "Filter by risk tier",
            options=["Critical", "High", "Medium", "Low", ""],
            default=["Critical", "High", "Medium", "Low", ""],
            format_func=lambda x: x if x else "No tier",
            key="gs_tier_filter",
        )
    with fc2:
        gs_resp_f = st.multiselect(
            "Filter by response",
            options=["No", "Partial", "N/A", "—"],
            default=["No", "Partial", "N/A", "—"],
            format_func=lambda x: "Unanswered" if x == "—" else x,
            key="gs_resp_filter",
        )
    with fc3:
        gs_dom_f = st.multiselect(
            "Filter by domain",
            options=sorted({i["domain_name"] for i in live_gaps}),
            default=sorted({i["domain_name"] for i in live_gaps}),
            key="gs_dom_filter",
        )

    filtered_gaps = [
        i for i in live_gaps
        if i["tier"] in gs_tier_f
        and i["norm"] in gs_resp_f
        and i["domain_name"] in gs_dom_f
    ]
    filtered_gaps.sort(key=lambda x: (TIER_ORDER.get(x["tier"], 4), x["domain"], x["key"]))

    st.caption(f"Showing {len(filtered_gaps)} of {total_gaps} gaps")

    # ── Export buttons ────────────────────────────────────────────────────────────
    export_df = pd.DataFrame([{
        "Key":        i["key"],
        "Domain":     i["domain_name"],
        "Question":   i["question"],
        "Response":   i["norm"],
        "Risk Tier":  i["tier"],
        "Remarks":    i["other"],
    } for i in filtered_gaps])

    ex1, ex2, ex3 = st.columns([1, 1, 1])
    with ex1:
        st.download_button(
            "⬇ Export gaps CSV",
            data=export_df.to_csv(index=False).encode(),
            file_name="tpcra_gap_summary.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with ex2:
        buf2 = io.BytesIO()
        with pd.ExcelWriter(buf2, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="Gap Summary")
        st.download_button(
            "⬇ Export gaps Excel",
            data=buf2.getvalue(),
            file_name="tpcra_gap_summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with ex3:
        pdf_bytes = generate_vendor_pdf(
            live_gaps=filtered_gaps,
            vendor_name=contact.get("vendor", ""),
            assessment_date=datetime.date.today().strftime("%d %B %Y"),
            total_questions=len(p2_items),
            gap_db=GAP_DB,
        )
        st.download_button(
            "⬇ Download vendor report PDF",
            data=pdf_bytes,
            file_name="tpcra_vendor_gap_report.pdf",
            mime="application/pdf",
            use_container_width=True,
            type="primary",
        )

    st.divider()

    # ── Stacked bar chart — gaps by domain × tier ────────────────────────────────
    chart_rows = []
    for letter, dom in domains.items():
        dom_gaps = [i for i in dom["items"] if i["norm"] in GAP_RESPONSES]
        for tier in ["Critical", "High", "Medium", "Low", ""]:
            cnt = sum(1 for i in dom_gaps if i["tier"] == tier)
            if cnt:
                chart_rows.append({
                    "Domain": letter,
                    "Tier": tier if tier else "No tier",
                    "Gaps": cnt,
                })

    if chart_rows:
        chart_df = pd.DataFrame(chart_rows)
        fig_gs = px.bar(
            chart_df, x="Domain", y="Gaps", color="Tier",
            color_discrete_map={
                "Critical": "#E24B4A",
                "High":     "#EF9F27",
                "Medium":   "#378ADD",
                "Low":      "#639922",
                "No tier":  "#B4B2A9",
            },
            category_orders={"Tier": ["Critical", "High", "Medium", "Low", "No tier"]},
            labels={"Gaps": "Gaps (No / N/A / Partial)", "Domain": "Domain"},
            height=260,
        )
        fig_gs.update_layout(
            margin=dict(l=0, r=0, t=10, b=0),
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            legend_title_text="Risk tier",
            font=dict(size=12),
            bargap=0.3,
        )
        fig_gs.update_xaxes(showgrid=False)
        fig_gs.update_yaxes(showgrid=True, gridcolor="rgba(0,0,0,0.07)")
        st.plotly_chart(fig_gs, use_container_width=True)

    st.divider()

    # ── Gaps grouped by domain ───────────────────────────────────────────────────
    TH_GS = (
        "padding:8px 10px;text-align:left;font-size:11px;font-weight:600;color:#6c757d;"
        "text-transform:uppercase;letter-spacing:0.05em;border-bottom:1px solid #e9ecef"
    )

    # Preserve domain order from DOMAIN_MAP
    domain_order = list(DOMAIN_MAP.keys())
    active_domain_letters = sorted(
        {i["domain"] for i in filtered_gaps},
        key=lambda x: domain_order.index(x) if x in domain_order else 99,
    )

    for dom_letter in active_domain_letters:
        dom_items = [i for i in filtered_gaps if i["domain"] == dom_letter]
        dom_name  = DOMAIN_MAP.get(dom_letter, dom_letter)

        # Sort within domain: Critical → High → Medium → Low → no tier, then by key
        dom_items.sort(key=lambda x: (TIER_ORDER.get(x["tier"], 4), x["key"]))

        n_no  = sum(1 for i in dom_items if i["norm"] == "No")
        n_pa  = sum(1 for i in dom_items if i["norm"] == "Partial")
        n_na  = sum(1 for i in dom_items if i["norm"] == "N/A")
        n_un  = sum(1 for i in dom_items if i["norm"] == "—")

        # Domain header
        st.markdown(
            f'<div style="display:flex;align-items:center;gap:10px;margin:12px 0 5px">'
            f'<span style="font-size:14px;font-weight:600;color:#333">'
            f'{dom_letter} — {dom_name}</span>'
            f'<span style="font-size:12px;color:#6c757d">'
            f'{len(dom_items)} gap{"s" if len(dom_items)!=1 else ""}'
            f'&nbsp;&nbsp;·&nbsp;&nbsp;No: <b>{n_no}</b>'
            f'&nbsp;&nbsp;Partial: <b>{n_pa}</b>'
            f'&nbsp;&nbsp;N/A: <b>{n_na}</b>'
            f'&nbsp;&nbsp;Unanswered: <b>{n_un}</b>'
            f'</span></div>',
            unsafe_allow_html=True,
        )

        # Gap table for this domain (no Domain column — redundant now)
        rows_html = ""
        for item in dom_items:
            resp_badge_html = (
                f'<span style="{GS_RESP_BADGE.get(item["norm"], GS_RESP_BADGE["—"])}">'
                f'{item["norm"] if item["norm"] != "—" else "Unanswered"}</span>'
            )
            tier_badge_html = (
                f'<span style="{GS_RISK_BADGE.get(item["tier"], GS_RISK_BADGE.get("Low", ""))}">'
                f'{item["tier"] if item["tier"] else "—"}</span>'
            )
            remarks_cell = (
                f'<div style="font-size:11px;color:#777;margin-top:4px;line-height:1.4">'
                f'{item["other"]}</div>'
                if item["other"] else ""
            )
            rows_html += (
                "<tr>"
                f'<td style="padding:8px 10px;border-bottom:1px solid #f0f0f0;font-size:11px;'
                f'color:#aaa;width:8%;vertical-align:top;white-space:nowrap">{item["key"]}</td>'
                f'<td style="padding:8px 10px;border-bottom:1px solid #f0f0f0;font-size:13px;'
                f'color:#333;line-height:1.5;width:67%;vertical-align:top">'
                f'{item["question"]}{remarks_cell}</td>'
                f'<td style="padding:8px 10px;border-bottom:1px solid #f0f0f0;width:12%;'
                f'vertical-align:top;text-align:center">{resp_badge_html}</td>'
                f'<td style="padding:8px 10px;border-bottom:1px solid #f0f0f0;width:13%;'
                f'vertical-align:top;text-align:center">{tier_badge_html}</td>'
                "</tr>"
            )

        st.markdown(
            '<div style="border:1px solid #e9ecef;border-radius:10px;overflow:hidden;margin-bottom:6px">'
            '<table style="width:100%;border-collapse:collapse;table-layout:fixed">'
            '<thead><tr style="background:#f8f9fa">'
            f'<th style="{TH_GS};width:8%">Ref</th>'
            f'<th style="{TH_GS};width:67%">Control / Question</th>'
            f'<th style="{TH_GS};width:12%;text-align:center">Response</th>'
            f'<th style="{TH_GS};width:13%;text-align:center">Tier</th>'
            "</tr></thead>"
            f"<tbody>{rows_html}</tbody>"
            "</table></div>",
            unsafe_allow_html=True,
        )

        # Per-domain recommendations
        dom_meta = next((d for d in GAP_DB if d["id"] == dom_letter), None)
        if dom_meta and dom_meta.get("recs"):
            with st.expander(f"Recommendations for {dom_letter} — {dom_name}", expanded=False):
                for idx, rec in enumerate(dom_meta["recs"], 1):
                    st.markdown(
                        f'<div style="padding:7px 12px;border-radius:7px;background:#f8f9fa;'
                        f'border-left:3px solid #185FA5;margin-bottom:5px;font-size:12px;'
                        f'color:#333;line-height:1.55">'
                        f'<span style="font-weight:600;color:#185FA5;margin-right:6px">{idx}.</span>{rec}'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

        st.markdown('<div style="margin-bottom:8px"></div>', unsafe_allow_html=True)


st.divider()
st.caption("Third-Party Cyber Risk Assessment Dashboard  ·  For internal use only")
